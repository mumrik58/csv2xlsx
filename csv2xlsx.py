#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Convert CSV to XLSX

This module convert file type from CSV to XLSX.

Example:
    $ python csv2xlsx.py test.csv

Attributes:

Todo:
    Nothing.
"""

import sys
import argparse
from logging import getLogger, StreamHandler, FileHandler, Formatter, DEBUG
from openpyxl import Workbook
from openpyxl.cell.cell import get_column_letter
import csv

if __name__ == '__main__':

    logger = getLogger(__name__)
    handler = StreamHandler(sys.stdout)
    handler.setFormatter(Formatter('%(asctime)s %(message)s'))
    handler.setLevel(DEBUG)
    logger.setLevel(DEBUG)
    logger.addHandler(handler)

    p = argparse.ArgumentParser()
    p.add_argument('input CSV file', type=str, help='input CSV file name')
    p.add_argument('-o', '--output', help='output XLSX file name', type=str)
    p.add_argument('-e', '--encoding', help='encoding type of input file', type=str, default=None)
    p.add_argument('-s', '--sheet-name', help='sheet name in output file', type=str)
    p.add_argument('--csv-field-size-limit', help='limit of CSV file size', type=int)

    args = p.parse_args()

    input_file = args.input
    output_file = '.'.join(input_file.split(
        '.')[0:-1]) + '.xlsx' if args.output is None else args.output
    wb = Workbook()

    # in case of using large CSV file.
    if args.csv_field_size_limit != 0:
        csv.field_size_limit(args.csv_field_size_limit)

    logger.info('input file: %s' % input_file)
    logger.info('input file: %s' % output_file)

    ws = wb.worksheets[0]

    if args.sheet_name is not None:
        ws.title = args.sheet_name

    with open(input_file, 'r', encoding=args.encoding) as f:
        dataReader = csv.reader(f)

        line_num = 0
        for line in dataReader:
            line_num += 1
        f.seek(0)

        for cnt_row, line in enumerate(dataReader):
            for cnt_col, word in enumerate(line):
                if word.__len__() != 0 and word[0] == '=':
                    word = '\'' + word
                ws[get_column_letter(cnt_col + 1).__str__() +
                   (cnt_row + 1).__str__()] = word
            if cnt_row % 1000 == 0:
                logger.debug(cnt_row.__str__() + '/' +
                             line_num.__str__() + ' lines were converted.')

    logger.debug('saving...')
    wb.save(filename=output_file)
    logger.info('done.')
